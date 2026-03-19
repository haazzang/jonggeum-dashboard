from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import re

import pandas as pd
from openpyxl import load_workbook

from dashboard_app.settings import DEFAULT_MANUAL_METRICS, INPUT_DIR, ROOT_DIR, WORKBOOK_SHEETS


KNOWN_ASSET_FILTERS = (
    "유동성자산현금성자산",
    "유동성자산기타유동성자산",
    "기업금융관련자산CP/STB",
    "기업금융관련자산채권",
    "기업금융관련자산기업대출",
    "기업금융관련자산출자",
    "부동산관련자산유동화증권",
    "부동산관련자산부동산",
    "기타자산CP/STB",
    "기타자산채권",
    "기타자산출자",
)


@dataclass(slots=True)
class SourceBundle:
    kind: str
    label: str
    modified_at: datetime
    workbook_path: Path | None = None
    deposits_csv_path: Path | None = None
    assets_csv_path: Path | None = None


@dataclass(slots=True)
class RawDataset:
    deposits: pd.DataFrame
    assets: pd.DataFrame
    source: SourceBundle
    base_date: datetime | None
    dashboard_seed_metrics: dict[str, float]


def load_raw_dataset() -> RawDataset:
    source = discover_latest_source()
    return load_raw_dataset_from_source(source)


def load_workbook_dataset(workbook_path: Path) -> RawDataset:
    source = SourceBundle(
        kind="workbook",
        label=f"기본 엑셀: {workbook_path.name}",
        workbook_path=workbook_path,
        modified_at=datetime.fromtimestamp(workbook_path.stat().st_mtime),
    )
    return load_raw_dataset_from_source(source)


def load_raw_dataset_from_source(source: SourceBundle) -> RawDataset:
    if source.kind == "workbook":
        deposits, assets, base_date, seed_metrics = _load_from_workbook(source.workbook_path)
    else:
        deposits = _read_csv_sheet(source.deposits_csv_path, "deposits")
        assets = _read_csv_sheet(source.assets_csv_path, "assets")
        base_date = _extract_base_date(deposits)
        seed_metrics = dict(DEFAULT_MANUAL_METRICS)

    return RawDataset(
        deposits=_normalize_deposits_frame(deposits),
        assets=_normalize_assets_frame(assets),
        source=source,
        base_date=base_date,
        dashboard_seed_metrics=seed_metrics,
    )


def discover_latest_source() -> SourceBundle:
    INPUT_DIR.mkdir(parents=True, exist_ok=True)
    candidates: list[SourceBundle] = []

    for workbook_path in _iter_workbooks(INPUT_DIR):
        candidates.append(
            SourceBundle(
                kind="workbook",
                label=f"업로드 엑셀: {workbook_path.name}",
                workbook_path=workbook_path,
                modified_at=datetime.fromtimestamp(workbook_path.stat().st_mtime),
            )
        )

    for directory in [INPUT_DIR, *[path for path in INPUT_DIR.rglob("*") if path.is_dir()]]:
        csv_pair = _find_csv_pair(directory)
        if csv_pair:
            deposits_path, assets_path = csv_pair
            candidates.append(
                SourceBundle(
                    kind="csv_bundle",
                    label=f"CSV 묶음: {directory.name or 'input'}",
                    deposits_csv_path=deposits_path,
                    assets_csv_path=assets_path,
                    modified_at=datetime.fromtimestamp(
                        max(deposits_path.stat().st_mtime, assets_path.stat().st_mtime)
                    ),
                )
            )

    for workbook_path in _iter_workbooks(ROOT_DIR):
        candidates.append(
            SourceBundle(
                kind="workbook",
                label=f"기본 엑셀: {workbook_path.name}",
                workbook_path=workbook_path,
                modified_at=datetime.fromtimestamp(workbook_path.stat().st_mtime),
            )
        )

    if not candidates:
        raise FileNotFoundError("읽을 수 있는 xlsx/csv 원본이 없습니다.")

    return max(candidates, key=lambda candidate: candidate.modified_at)


def _iter_workbooks(directory: Path) -> list[Path]:
    return [
        path
        for path in directory.glob("*.xlsx")
        if path.is_file() and not path.name.startswith("~$")
    ]


def _find_csv_pair(directory: Path) -> tuple[Path, Path] | None:
    csv_files = [path for path in directory.glob("*.csv") if path.is_file()]
    if not csv_files:
        return None

    deposits_path = None
    assets_path = None
    for path in csv_files:
        name = path.stem.lower()
        if "수신잔고" in path.stem or "deposit" in name:
            deposits_path = path
        elif "운용자산" in path.stem or "asset" in name:
            assets_path = path

    if deposits_path and assets_path:
        return deposits_path, assets_path
    return None


def _load_from_workbook(workbook_path: Path | None) -> tuple[pd.DataFrame, pd.DataFrame, datetime | None, dict[str, float]]:
    if workbook_path is None:
        raise FileNotFoundError("엑셀 경로가 없습니다.")

    workbook = load_workbook(workbook_path, data_only=True)
    deposits_sheet = _find_sheet(workbook, WORKBOOK_SHEETS["deposits"])
    assets_sheet = _find_sheet(workbook, WORKBOOK_SHEETS["assets"])
    dashboard_sheet = _find_sheet(workbook, WORKBOOK_SHEETS["dashboard"], required=False)

    deposits = _worksheet_to_frame(deposits_sheet, header_row=4, data_start_row=6)
    assets = _worksheet_to_frame(assets_sheet, header_row=4, data_start_row=5)

    base_date = None
    seed_metrics = dict(DEFAULT_MANUAL_METRICS)
    if dashboard_sheet is not None:
        base_date = dashboard_sheet["T3"].value
        seed_metrics = {
            "venture_capital_issued_note_ratio": float(dashboard_sheet["Q8"].value or 0.0),
            "venture_capital_company_wide_quarter_ratio": float(dashboard_sheet["R8"].value or 0.0),
            "liquidity_ratio_1m": float(dashboard_sheet["S8"].value or 0.0),
            "liquidity_ratio_3m": float(dashboard_sheet["T8"].value or 0.0),
        }

    if base_date is None:
        base_date = _extract_base_date(deposits)

    return deposits, assets, base_date, seed_metrics


def _find_sheet(workbook, candidates: tuple[str, ...], required: bool = True):
    for name in candidates:
        if name in workbook.sheetnames:
            return workbook[name]
    if required:
        raise KeyError(f"시트를 찾을 수 없습니다: {candidates}")
    return None


def _worksheet_to_frame(worksheet, header_row: int, data_start_row: int) -> pd.DataFrame:
    rows = list(worksheet.values)
    header_index = header_row - 1
    data_index = data_start_row - 1
    headers = _deduplicate_headers(rows[header_index])
    frame = pd.DataFrame(rows[data_index:], columns=headers)
    return frame.dropna(how="all").reset_index(drop=True)


def _read_csv_sheet(csv_path: Path | None, dataset_kind: str) -> pd.DataFrame:
    if csv_path is None:
        raise FileNotFoundError("CSV 경로가 없습니다.")

    raw = _read_csv_with_fallbacks(csv_path)
    keywords = {
        "deposits": ("상품구분", "발행잔고"),
        "assets": ("자산명", "BS금액"),
    }
    header_idx = _detect_header_row(raw, keywords[dataset_kind])
    data_start_idx = header_idx + (2 if dataset_kind == "deposits" else 1)
    headers = _deduplicate_headers(raw.iloc[header_idx].tolist())
    frame = raw.iloc[data_start_idx:].copy()
    frame.columns = headers
    return frame.dropna(how="all").reset_index(drop=True)


def _read_csv_with_fallbacks(csv_path: Path) -> pd.DataFrame:
    encodings = ("utf-8-sig", "cp949", "euc-kr")
    last_error = None
    for encoding in encodings:
        try:
            return pd.read_csv(csv_path, header=None, dtype=object, encoding=encoding)
        except UnicodeDecodeError as error:
            last_error = error
    raise last_error if last_error else ValueError(f"CSV를 읽을 수 없습니다: {csv_path}")


def _detect_header_row(raw: pd.DataFrame, keywords: tuple[str, str]) -> int:
    for idx in range(min(len(raw), 10)):
        row_text = {str(value).strip() for value in raw.iloc[idx].tolist() if value is not None}
        if all(keyword in row_text for keyword in keywords):
            return idx
    raise ValueError(f"CSV 헤더 행을 찾을 수 없습니다. 필요 키워드: {keywords}")


def _deduplicate_headers(headers) -> list[str]:
    seen: dict[str, int] = {}
    cleaned_headers: list[str] = []
    for index, header in enumerate(headers):
        name = str(header).strip() if header is not None else f"__unnamed_{index + 1}"
        if not name:
            name = f"__unnamed_{index + 1}"
        duplicate_count = seen.get(name, 0)
        seen[name] = duplicate_count + 1
        if duplicate_count:
            name = f"{name}_{duplicate_count + 1}"
        cleaned_headers.append(name)
    return cleaned_headers


def _normalize_deposits_frame(frame: pd.DataFrame) -> pd.DataFrame:
    normalized = frame.copy()
    normalized.columns = [str(column).strip() for column in normalized.columns]

    filter1_series = _get_series(normalized, "필터1")
    filter2_series = _get_series(normalized, "필터2")
    product_series = _get_series(normalized, "상품구분")

    derived_filter1 = product_series.apply(_derive_deposit_filter1)
    normalized["dashboard_filter1"] = filter1_series.combine_first(derived_filter1)
    normalized["dashboard_filter2"] = filter2_series.combine_first(
        normalized["dashboard_filter1"].apply(_derive_deposit_filter2)
    )

    normalized["dashboard_product_name"] = product_series
    normalized["dashboard_balance"] = pd.to_numeric(
        _coalesce_series(normalized, "발행잔고", "발행어음잔고"),
        errors="coerce",
    ).fillna(0.0)
    normalized["dashboard_rate"] = pd.to_numeric(
        _coalesce_series(normalized, "조달금리", "조달금리(평균)"),
        errors="coerce",
    ).fillna(0.0)
    normalized["dashboard_date"] = pd.to_datetime(_get_series(normalized, "기준일자"), errors="coerce")
    normalized["dashboard_is_summary"] = normalized["dashboard_product_name"].astype(str).str.contains(
        "합계|총 금액", na=False
    )
    return normalized


def _normalize_assets_frame(frame: pd.DataFrame) -> pd.DataFrame:
    normalized = frame.copy()
    normalized.columns = [str(column).strip() for column in normalized.columns]

    filter_series = _get_series(normalized, "필터")
    asset_name_series = _get_series(normalized, "자산명")

    derived_filter = asset_name_series.apply(_derive_asset_filter)
    normalized["dashboard_filter"] = filter_series.combine_first(derived_filter)
    normalized["dashboard_asset_name"] = asset_name_series
    normalized["dashboard_balance"] = pd.to_numeric(_get_series(normalized, "BS금액"), errors="coerce").fillna(
        0.0
    )
    normalized["dashboard_rate"] = pd.to_numeric(
        _coalesce_series(normalized, "운용금리(%)", "운용금리"),
        errors="coerce",
    ).fillna(0.0)
    normalized["dashboard_duration"] = pd.to_numeric(
        _coalesce_series(normalized, "듀레이션", "Duration"),
        errors="coerce",
    ).fillna(0.0)
    return normalized


def _derive_deposit_filter1(value) -> str | None:
    if value is None or pd.isna(value):
        return None
    text = str(value).strip()
    if not text or "합계" in text or text == "총 금액":
        return None

    parts = text.rsplit("_", 1)
    category = parts[-1] if len(parts) > 1 else text
    return category.strip()


def _derive_deposit_filter2(value) -> str | None:
    if value is None or pd.isna(value):
        return None
    text = str(value).strip()
    if text.startswith("개인") or text.startswith("법인"):
        return text[2:]
    return text


def _derive_asset_filter(value) -> str | None:
    if value is None or pd.isna(value):
        return None
    text = re.sub(r"\d+$", "", str(value).strip())
    for known_filter in KNOWN_ASSET_FILTERS:
        if text.startswith(known_filter):
            return known_filter
    return text or None


def _extract_base_date(deposits: pd.DataFrame) -> datetime | None:
    if "기준일자" in deposits:
        dates = pd.to_datetime(deposits["기준일자"], errors="coerce").dropna()
        if not dates.empty:
            return dates.iloc[0].to_pydatetime()
    if "dashboard_date" in deposits:
        dates = pd.to_datetime(deposits["dashboard_date"], errors="coerce").dropna()
        if not dates.empty:
            return dates.iloc[0].to_pydatetime()
    return None


def _get_series(frame: pd.DataFrame, column_name: str) -> pd.Series:
    if column_name in frame.columns:
        return frame[column_name]
    return pd.Series(index=frame.index, dtype=object)


def _coalesce_series(frame: pd.DataFrame, *column_names: str) -> pd.Series:
    result = pd.Series([None] * len(frame), index=frame.index, dtype=object)
    for column_name in column_names:
        current = _get_series(frame, column_name)
        mask = result.isna() & current.notna()
        result.loc[mask] = current.loc[mask]
    return result
