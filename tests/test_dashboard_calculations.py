from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from dashboard_app.calculations import build_dashboard
from dashboard_app.data_loader import RawDataset, SourceBundle, load_workbook_dataset
from dashboard_app.manual_metrics import ManualMetrics


ROOT = Path(__file__).resolve().parent.parent
SAMPLE_PATH = ROOT / "대시보드 AI 기초자료 만들기_sample.xlsx"


def test_sample_workbook_matches_core_dashboard_metrics():
    raw_dataset = load_workbook_dataset(SAMPLE_PATH)
    workbook = load_workbook(SAMPLE_PATH, data_only=True)
    dashboard_sheet = workbook["DASHBOARD"]

    manual_metrics = ManualMetrics(
        venture_capital_issued_note_ratio=float(dashboard_sheet["Q8"].value or 0.0),
        venture_capital_company_wide_quarter_ratio=float(dashboard_sheet["R8"].value or 0.0),
        liquidity_ratio_1m=float(dashboard_sheet["S8"].value or 0.0),
        liquidity_ratio_3m=float(dashboard_sheet["T8"].value or 0.0),
    )
    dashboard = build_dashboard(raw_dataset, manual_metrics)
    summary = dashboard["business_summary"]
    asset_sections = dashboard["asset_section"]["sections"]
    deposit_section = dashboard["deposit_section"]

    assert summary["deposits_balance"] == _approx(dashboard_sheet["E8"].value)
    assert summary["assets_yield"] == _approx(dashboard_sheet["G8"].value)
    assert summary["funding_rate"] == _approx(dashboard_sheet["H8"].value)
    assert summary["operating_margin"] == _approx(dashboard_sheet["J8"].value)
    assert dashboard["asset_section"]["total_balance"] == _approx(dashboard_sheet["F14"].value)
    assert deposit_section["total_balance"] == _approx(dashboard_sheet["P14"].value)
    assert deposit_section["personal_total"] == _approx(dashboard_sheet["Q14"].value)
    assert deposit_section["corporate_total"] == _approx(dashboard_sheet["R14"].value)
    assert deposit_section["total_rate"] == _approx(dashboard_sheet["S14"].value)

    labels = {section["label"]: section for section in asset_sections}
    assert labels["유동성자산"]["balance"] == _approx(dashboard_sheet["F15"].value)
    assert labels["기업금융관련자산"]["balance"] == _approx(dashboard_sheet["F18"].value)
    assert labels["부동산관련자산"]["balance"] == _approx(dashboard_sheet["F23"].value)
    assert labels["기타자산"]["balance"] == _approx(dashboard_sheet["F26"].value)
    assert dashboard["trend_chart"]["available"] is False


def test_daily_trend_chart_payload_is_exposed_for_dynamic_chart():
    raw_dataset = load_workbook_dataset(SAMPLE_PATH)
    trend_frame = pd.DataFrame(
        [
            {
                "series_key": "운용자산",
                "series_label": "운용자산",
                "series_path": "운용자산",
                "level_1": "운용자산",
                "level_2": "",
                "level_3": "",
                "level_4": "",
                "depth": 1,
                "2026-01-01": 325_000_000_000,
                "2026-01-02": 330_000_000_000,
                "2026-01-03": 333_000_000_000,
            },
            {
                "series_key": "수신잔고",
                "series_label": "수신잔고",
                "series_path": "수신잔고",
                "level_1": "수신잔고",
                "level_2": "",
                "level_3": "",
                "level_4": "",
                "depth": 1,
                "2026-01-01": 335_000_000_000,
                "2026-01-02": 340_000_000_000,
                "2026-01-03": 338_000_000_000,
            },
            {
                "series_key": "수신잔고 > 약정형",
                "series_label": "약정형",
                "series_path": "수신잔고 > 약정형",
                "level_1": "수신잔고",
                "level_2": "약정형",
                "level_3": "",
                "level_4": "",
                "depth": 2,
                "2026-01-01": 270_000_000_000,
                "2026-01-02": 272_000_000_000,
                "2026-01-03": 273_000_000_000,
            },
        ]
    )
    enriched_dataset = RawDataset(
        deposits=raw_dataset.deposits,
        assets=raw_dataset.assets,
        daily_trend=trend_frame,
        source=SourceBundle(
            kind="workbook",
            label="synthetic trend",
            modified_at=datetime(2026, 1, 3, 9, 0, 0),
            workbook_path=SAMPLE_PATH,
        ),
        base_date=raw_dataset.base_date,
        dashboard_seed_metrics=raw_dataset.dashboard_seed_metrics,
    )

    dashboard = build_dashboard(
        enriched_dataset,
        ManualMetrics(
            venture_capital_issued_note_ratio=0.0,
            venture_capital_company_wide_quarter_ratio=0.8209,
            liquidity_ratio_1m=6.0407,
            liquidity_ratio_3m=5.5094,
        ),
    )

    trend_chart = dashboard["trend_chart"]

    assert trend_chart["available"] is True
    assert trend_chart["default_key"] == "수신잔고"
    assert trend_chart["labels"] == ["2026-01-01", "2026-01-02", "2026-01-03"]
    assert trend_chart["series_count"] == 3
    assert trend_chart["featured_keys"][:2] == ["수신잔고", "운용자산"]
    assert trend_chart["series"]["운용자산"]["values"] == pytest.approx([3250.0, 3300.0, 3330.0])
    assert trend_chart["series"]["수신잔고"]["path"] == "수신잔고"


def _approx(value):
    return pytest.approx(float(value), rel=1e-9, abs=1e-9)
